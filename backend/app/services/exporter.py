from __future__ import annotations

import json
import shutil
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..database import db_cursor, get_extracted_data, now_iso
from ..schemas import ExportFilters
from .formatters import extension_for_path, format_invoice_date
from .invoice_extractor import normalize_prompt_fields


SUMMARY_EXCEL_HEADERS = [
    "Archive Number",
    "Expense Type",
    "Vendor Code",
    "Vendor Name",
    "PO",
    "Invoice Number",
    "Invoice Date",
    "Amount",
    "Invoice Category",
]

DETAIL_SYSTEM_COLUMNS = [
    {"key": "archive_number", "label": "Archive Number", "source": "scalar", "row_mode": "repeat", "type": "string"},
    {"key": "expense_type", "label": "Expense Type", "source": "scalar", "row_mode": "repeat", "type": "string"},
    {"key": "vendor_code", "label": "Vendor Code", "source": "scalar", "row_mode": "repeat", "type": "string"},
    {"key": "vendor_name", "label": "Vendor Name", "source": "scalar", "row_mode": "repeat", "type": "string"},
    {"key": "po_number", "label": "PO", "source": "scalar", "row_mode": "repeat", "type": "string"},
    {"key": "invoice_number", "label": "Invoice Number", "source": "scalar", "row_mode": "repeat", "type": "string"},
    {"key": "invoice_date", "label": "Invoice Date", "source": "scalar", "row_mode": "repeat", "type": "string"},
    {"key": "total_amount", "label": "Amount", "source": "scalar", "row_mode": "repeat", "type": "value"},
    {"key": "invoice_category", "label": "Invoice Category", "source": "scalar", "row_mode": "repeat", "type": "string"},
]

SYSTEM_DEDUPE_KEYS = {
    "archive_number",
    "expense_type",
    "vendor_code",
    "supplier_code",
    "vendor_name",
    "po_number",
    "po",
    "customer_po",
    "invoice_number",
    "invoice_date",
    "total_amount",
    "invoice_category",
}

INVALID_SHEET_CHARS = set("[]:*?/\\")


def _matches_filters(invoice: dict[str, Any], filters: ExportFilters) -> bool:
    mode = filters.mode
    if mode == "all":
        return True
    if mode == "day":
        day = filters.day or ""
        return bool(day) and invoice.get("invoice_date_iso") == day
    if mode == "range":
        if filters.confirmed_from or filters.confirmed_to:
            confirmed_at = invoice.get("confirmed_at") or ""
            if not confirmed_at:
                return False
            if filters.confirmed_from and confirmed_at < filters.confirmed_from:
                return False
            if filters.confirmed_to and confirmed_at > filters.confirmed_to:
                return False
        else:
            date_value = invoice.get("invoice_date_iso") or ""
            if not date_value:
                return False
            if filters.date_from and date_value < filters.date_from:
                return False
            if filters.date_to and date_value > filters.date_to:
                return False
        return True
    if mode == "supplier":
        needle = (filters.supplier or "").strip().lower()
        if not needle:
            return True
        haystack = f"{invoice.get('vendor_code', '')} {invoice.get('vendor_name', '')}".lower()
        return needle in haystack
    if mode == "category":
        category = (filters.category or "").strip()
        if not category:
            return True
        return str(invoice.get("invoice_category") or "").strip() == category
    return True


def _load_confirmed(filters: ExportFilters, invoice_ids: list[int] | None = None) -> list[dict[str, Any]]:
    if invoice_ids:
        unique_ids = list(dict.fromkeys(invoice_ids))
        placeholders = ",".join("?" for _ in unique_ids)
        with db_cursor() as cur:
            rows = cur.execute(
                f"""
                SELECT * FROM invoices
                WHERE status = 'confirmed'
                    AND id IN ({placeholders})
                """,
                unique_ids,
            ).fetchall()
        invoices_by_id = {row["id"]: dict(row) for row in rows}
        return [invoices_by_id[invoice_id] for invoice_id in unique_ids if invoice_id in invoices_by_id]

    with db_cursor() as cur:
        rows = cur.execute(
            """
            SELECT * FROM invoices
            WHERE status = 'confirmed'
            ORDER BY confirmed_at ASC, id ASC
            """
        ).fetchall()
    invoices = [dict(row) for row in rows]
    return [invoice for invoice in invoices if _matches_filters(invoice, filters)]


def _write_excel(
    path: Path,
    summary_rows: list[dict[str, Any]],
    tag_groups: dict[str, list[dict[str, Any]]],
) -> None:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    _write_plain_sheet(summary_ws, SUMMARY_EXCEL_HEADERS, summary_rows, table_name="InvoiceSummary")

    used_titles = {"Summary"}
    for tag_name, records in tag_groups.items():
        title = _safe_sheet_title(tag_name, used_titles)
        used_titles.add(title)
        ws = wb.create_sheet(title)
        fields, settings = _tag_export_config(tag_name)
        columns = _export_columns_for_tag(fields, settings)
        _write_detail_sheet(ws, columns, records)

    wb.save(path)


def _write_plain_sheet(ws: Any, headers: list[str], rows: list[dict[str, Any]], table_name: str) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    _style_sheet(ws, headers)

    if rows and headers:
        last_column = get_column_letter(len(headers))
        table_ref = f"A1:{last_column}{len(rows) + 1}"
        table = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)


def _write_detail_sheet(ws: Any, columns: list[dict[str, Any]], records: list[dict[str, Any]]) -> None:
    headers = [_column_label(column) for column in columns]
    ws.append(headers)

    next_row = 2
    merge_ranges: list[tuple[int, int, int]] = []
    for record in records:
        rows, merges = _detail_rows_for_invoice(record, columns, next_row)
        for row in rows:
            ws.append(row)
        merge_ranges.extend(merges)
        next_row += len(rows)

    for start_row, end_row, col_index in merge_ranges:
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=col_index, end_row=end_row, end_column=col_index)

    _style_sheet(ws, headers)
    if records and headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def _style_sheet(ws: Any, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="00A6D6")
    cell_border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = cell_border
            if isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = "#,##0.00"

    for index, header in enumerate(headers, start=1):
        values = [str(header)]
        for row_index in range(2, min(ws.max_row, 80) + 1):
            value = ws.cell(row=row_index, column=index).value
            if value not in (None, ""):
                values.append(str(value))
        width = min(max(max((len(value) for value in values), default=10) + 3, 12), 36)
        ws.column_dimensions[get_column_letter(index)].width = width


def _safe_sheet_title(raw_title: str, used_titles: set[str]) -> str:
    title = "".join("_" if char in INVALID_SHEET_CHARS else char for char in str(raw_title or "").strip())
    title = title[:31].strip() or "default"
    candidate = title
    suffix = 2
    while candidate in used_titles:
        tail = f"_{suffix}"
        candidate = f"{title[:31 - len(tail)]}{tail}"
        suffix += 1
    return candidate


def _numberish(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return ""
    clean = text.replace("$", "").replace(",", "").strip()
    match = clean.split()[0] if clean.split() else clean
    try:
        return float(match)
    except ValueError:
        return text


def _first_present(data: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        value = data.get(key)
        if value not in (None, ""):
            return value
    return ""


def _array_items_from_data(data: dict[str, Any], array_key: str) -> list[dict[str, Any]]:
    raw = data.get(array_key)
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            return []
        try:
            raw = json.loads(text)
        except json.JSONDecodeError:
            return []
    if not isinstance(raw, list):
        return []
    items: list[dict[str, Any]] = []
    for item in raw:
        if isinstance(item, dict):
            items.append(item)
    return items


def _tag_export_config(tag_name: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    tag = str(tag_name or "").strip() or "default"
    with db_cursor() as cur:
        row = cur.execute(
            "SELECT fields_json, export_settings_json FROM schemes WHERE name = ?",
            (tag,),
        ).fetchone()
        if not row and tag != "default":
            row = cur.execute(
                "SELECT fields_json, export_settings_json FROM schemes WHERE name = 'default'",
            ).fetchone()
    if not row:
        return [], {}
    try:
        raw = json.loads(row["fields_json"] or "[]")
    except json.JSONDecodeError:
        raw = []
    try:
        settings = json.loads(row["export_settings_json"] or "{}")
    except json.JSONDecodeError:
        settings = {}
    return normalize_prompt_fields(raw), settings if isinstance(settings, dict) else {}


def _export_columns_for_tag(fields: list[dict[str, Any]], settings: dict[str, Any]) -> list[dict[str, Any]]:
    raw_columns = settings.get("columns") if settings.get("custom") else None
    if isinstance(raw_columns, list):
        columns: list[dict[str, Any]] = []
        seen: set[str] = set()
        for item in raw_columns:
            if not isinstance(item, dict):
                continue
            column = _normalize_export_column(item)
            if not column:
                continue
            seen.add(str(column["key"]).lower())
            if column.get("enabled", True):
                columns.append(column)
        for column in _default_export_columns_for_fields(fields):
            key = str(column.get("key") or "").lower()
            if key in seen:
                continue
            seen.add(key)
            columns.append(column)
        return columns

    return _default_export_columns_for_fields(fields)


def _default_export_columns_for_fields(fields: list[dict[str, Any]]) -> list[dict[str, Any]]:
    columns: list[dict[str, Any]] = []
    columns.extend(dict(item, enabled=True) for item in DETAIL_SYSTEM_COLUMNS)
    for field in fields:
        field_type = str(field.get("type") or "string")
        key = str(field.get("key") or "").strip()
        if not key:
            continue
        if field_type == "array":
            for child in field.get("children") or []:
                child_key = str(child.get("key") or "").strip()
                if not child_key:
                    continue
                columns.append(
                    {
                        "key": f"{key}.{child_key}",
                        "label": _excel_header_for_field(child_key),
                        "enabled": True,
                        "source": "array_child",
                        "row_mode": "repeat",
                        "array_key": key,
                        "child_key": child_key,
                        "type": str(child.get("type") or "string"),
                    }
                )
            continue
        if key.strip().lower() in SYSTEM_DEDUPE_KEYS:
            continue
        columns.append(
            {
                "key": key,
                "label": _excel_header_for_field(key),
                "enabled": True,
                "source": "scalar",
                "row_mode": "repeat",
                "array_key": "",
                "child_key": "",
                "type": field_type,
            }
        )
    return columns


def _normalize_export_column(item: dict[str, Any]) -> dict[str, Any] | None:
    key = str(item.get("key") or "").strip()
    if not key:
        return None
    source = str(item.get("source") or "scalar")
    if source not in {"scalar", "array_child"}:
        source = "scalar"
    row_mode = str(item.get("row_mode") or "repeat")
    if row_mode not in {"merge", "repeat", "split_even"}:
        row_mode = "repeat"
    array_key = str(item.get("array_key") or "").strip()
    child_key = str(item.get("child_key") or "").strip()
    if source == "array_child":
        if not array_key and "." in key:
            array_key, child_key = key.split(".", 1)
        if not array_key or not child_key:
            return None
    return {
        "key": key,
        "label": str(item.get("label") or "").strip() or _excel_header_for_field(child_key or key),
        "enabled": item.get("enabled", True) is not False,
        "source": source,
        "row_mode": row_mode,
        "array_key": array_key,
        "child_key": child_key,
        "type": str(item.get("type") or "string"),
    }


def _excel_header_for_field(key: str) -> str:
    normalized = key.strip().lower()
    labels = {
        "archive_number": "Archive Number",
        "expense_type": "Expense Type",
        "vendor_code": "Vendor Code",
        "supplier_code": "Vendor Code",
        "vendor_name": "Vendor Name",
        "invoice_category": "Invoice Category",
        "po_number": "PO",
        "po": "PO",
        "customer_po": "PO",
        "invoice_number": "Invoice Number",
        "invoice_date": "Invoice Date",
        "bol_number": "BOL",
        "bill_of_lading": "BOL",
        "weight": "Weight",
        "weight_tn": "Weight",
        "qty_tn": "Weight",
        "qty_priced": "Weight",
        "unit_price": "Unit Price",
        "commodity_amount": "Commodity Amount",
        "freight_amount": "Freight",
        "tax_amount": "Tax Amount",
        "line_amount": "Amount",
        "amount": "Amount",
        "total_amount": "Amount",
    }
    return labels.get(normalized, key)


def _summary_row(invoice: dict[str, Any], archive_number: str) -> dict[str, Any]:
    return {
        "Archive Number": archive_number,
        "Expense Type": invoice.get("expense_type") or "",
        "Vendor Code": invoice.get("vendor_code") or "",
        "Vendor Name": invoice.get("vendor_name") or "",
        "PO": invoice.get("po_number") or "",
        "Invoice Number": invoice.get("invoice_number") or "",
        "Invoice Date": format_invoice_date(invoice.get("invoice_date_iso") or invoice.get("invoice_date") or ""),
        "Amount": invoice.get("total_amount") or 0,
        "Invoice Category": invoice.get("invoice_category") or "",
    }


def _detail_rows_for_invoice(
    record: dict[str, Any],
    columns: list[dict[str, Any]],
    start_row: int,
) -> tuple[list[list[Any]], list[tuple[int, int, int]]]:
    invoice = record["invoice"]
    data = record["data"]
    archive_number = record["archive_number"]
    array_key = _primary_array_key(columns)
    array_items = _array_items_from_data(data, array_key) if array_key else []
    row_count = max(1, len(array_items))
    rows = [["" for _ in columns] for _ in range(row_count)]
    merge_ranges: list[tuple[int, int, int]] = []

    for col_index, column in enumerate(columns, start=1):
        source = column.get("source")
        if source == "array_child":
            child_key = str(column.get("child_key") or "").strip()
            for row_index in range(row_count):
                item = array_items[row_index] if row_index < len(array_items) else {}
                value = _case_insensitive_get(item, child_key)
                rows[row_index][col_index - 1] = _coerce_export_value(value, column.get("type"))
            continue

        value = _scalar_value(invoice, data, archive_number, str(column.get("key") or ""))
        row_mode = str(column.get("row_mode") or "repeat")
        if row_mode == "split_even":
            split_values = _split_even(value, row_count)
            for row_index, split_value in enumerate(split_values):
                rows[row_index][col_index - 1] = split_value
            continue
        if row_mode == "repeat":
            cell_value = _coerce_export_value(value, column.get("type"))
            for row_index in range(row_count):
                rows[row_index][col_index - 1] = cell_value
            continue

        rows[0][col_index - 1] = _coerce_export_value(value, column.get("type"))
        if row_count > 1:
            merge_ranges.append((start_row, start_row + row_count - 1, col_index))

    return rows, merge_ranges


def _primary_array_key(columns: list[dict[str, Any]]) -> str:
    for column in columns:
        if column.get("source") == "array_child":
            array_key = str(column.get("array_key") or "").strip()
            if array_key:
                return array_key
    return ""


def _column_label(column: dict[str, Any]) -> str:
    return str(column.get("label") or "").strip() or _excel_header_for_field(str(column.get("key") or ""))


def _case_insensitive_get(data: dict[str, Any], key: str) -> Any:
    if key in data:
        return data.get(key)
    normalized = key.strip().lower()
    for item_key, value in data.items():
        if str(item_key).strip().lower() == normalized:
            return value
    return ""


def _scalar_value(invoice: dict[str, Any], data: dict[str, Any], archive_number: str, key: str) -> Any:
    normalized = key.strip().lower()
    if normalized == "archive_number":
        return archive_number
    if normalized == "expense_type":
        return invoice.get("expense_type") or _case_insensitive_get(data, key)
    if normalized in {"vendor_code", "supplier_code"}:
        return invoice.get("vendor_code") or _first_present_ci(data, ["vendor_code", "supplier_code", "Vendor Code"])
    if normalized == "vendor_name":
        return invoice.get("vendor_name") or _first_present_ci(data, ["vendor_name", "Vendor Name"])
    if normalized in {"po_number", "po", "customer_po"}:
        return invoice.get("po_number") or _first_present_ci(data, ["PO_number", "po_number", "customer_po", "PO"])
    if normalized == "invoice_number":
        return invoice.get("invoice_number") or _first_present_ci(data, ["invoice_number", "Invoice Number"])
    if normalized == "invoice_date":
        return format_invoice_date(
            invoice.get("invoice_date_iso")
            or invoice.get("invoice_date")
            or _first_present_ci(data, ["invoice_date", "Invoice Date"])
        )
    if normalized in {"amount", "total_amount"}:
        return invoice.get("total_amount") or _first_present_ci(data, ["total_amount", "amount", "Amount"])
    if normalized == "invoice_category":
        return invoice.get("invoice_category") or _first_present_ci(data, ["invoice_category", "Invoice Category"])
    if normalized == "prompt_tag":
        return data.get("prompt_tag") or "default"
    value = _case_insensitive_get(data, key)
    if value not in (None, ""):
        return value
    return invoice.get(key, "")


def _first_present_ci(data: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        value = _case_insensitive_get(data, key)
        if value not in (None, ""):
            return value
    return ""


def _coerce_export_value(value: Any, value_type: Any) -> Any:
    if str(value_type or "").lower() == "value":
        return _numberish(value)
    return "" if value is None else value


def _decimal_amount(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    clean = str(value).replace("$", "").replace(",", "").strip()
    try:
        return Decimal(clean).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def _split_even(value: Any, parts: int) -> list[Decimal]:
    count = max(1, parts)
    amount = _decimal_amount(value)
    cents = int((abs(amount) * 100).to_integral_value(rounding=ROUND_HALF_UP))
    base, remainder = divmod(cents, count)
    sign = -1 if amount < 0 else 1
    values = []
    for index in range(count):
        item_cents = base + (1 if index < remainder else 0)
        values.append((Decimal(sign * item_cents) / Decimal(100)).quantize(Decimal("0.01")))
    return values


def export_confirmed(
    destination_dir: str,
    prefix: str,
    start_number_text: str,
    filters: ExportFilters,
    invoice_ids: list[int] | None = None,
    create_new_folder: bool = False,
) -> dict[str, Any]:
    base_destination = Path(destination_dir).expanduser()
    base_destination.mkdir(parents=True, exist_ok=True)
    if not base_destination.is_dir():
        raise ValueError("Export destination is not a folder")

    prefix = prefix.strip()
    if not prefix:
        raise ValueError("Archive number prefix cannot be blank")
    if not start_number_text.isdigit():
        raise ValueError("Start number must be numeric")

    batch_id = uuid.uuid4().hex
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destination = base_destination
    if create_new_folder:
        destination = base_destination / f"invoice_export_{timestamp}_{batch_id[:6]}"

    start_number = int(start_number_text)
    width = max(4, len(start_number_text))
    invoices = _load_confirmed(filters, invoice_ids)
    if not invoices:
        raise ValueError("No matching confirmed invoices")

    planned: list[dict[str, Any]] = []
    for offset, invoice in enumerate(invoices):
        archive_number = f"{prefix}{start_number + offset:0{width}d}"
        extension = extension_for_path(invoice["file_path"])
        exported_filename = f"{archive_number}{extension}"
        exported_path = destination / exported_filename
        if exported_path.exists():
            raise FileExistsError(f"Target file already exists: {exported_path}")
        source = Path(invoice["file_path"])
        if not source.exists():
            raise FileNotFoundError(f"Confirmed file does not exist: {source}")
        planned.append(
            {
                "invoice": invoice,
                "archive_number": archive_number,
                "exported_filename": exported_filename,
                "exported_path": exported_path,
            }
        )

    if create_new_folder:
        destination.mkdir(parents=True, exist_ok=False)

    excel_path = destination / f"invoice_export_{timestamp}_{batch_id[:6]}.xlsx"

    summary_rows: list[dict[str, Any]] = []
    tag_groups: dict[str, list[dict[str, Any]]] = {}
    exported_files: list[dict[str, str]] = []
    created_at = now_iso()
    for item in planned:
        invoice = item["invoice"]
        source = Path(invoice["file_path"])
        shutil.copyfile(source, item["exported_path"])
        data = get_extracted_data(int(invoice["id"]))
        tag_name = str(data.get("prompt_tag") or "default").strip() or "default"
        record = {
            "invoice": invoice,
            "data": data,
            "archive_number": item["archive_number"],
        }
        summary_rows.append(_summary_row(invoice, item["archive_number"]))
        tag_groups.setdefault(tag_name, []).append(record)
        exported_files.append(
            {
                "invoice_id": str(invoice["id"]),
                "archive_number": item["archive_number"],
                "exported_filename": item["exported_filename"],
                "exported_path": str(item["exported_path"]),
            }
        )

    _write_excel(excel_path, summary_rows, tag_groups)

    with db_cursor() as cur:
        cur.execute(
            """
            INSERT INTO export_batches(
                id, destination_dir, prefix, start_number, number_width,
                filters_json, item_count, excel_path, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                batch_id,
                str(destination),
                prefix,
                start_number,
                width,
                json.dumps(filters.model_dump(), ensure_ascii=False),
                len(planned),
                str(excel_path),
                created_at,
            ),
        )
        for item in planned:
            invoice = item["invoice"]
            cur.execute(
                """
                INSERT INTO export_items(
                    batch_id, invoice_id, archive_number, exported_filename,
                    exported_path, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    invoice["id"],
                    item["archive_number"],
                    item["exported_filename"],
                    str(item["exported_path"]),
                    created_at,
                ),
            )

    return {
        "batch_id": batch_id,
        "item_count": len(planned),
        "destination_dir": str(destination),
        "excel_path": str(excel_path),
        "files": exported_files,
    }
